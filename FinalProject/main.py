import sqlite3
from datetime import datetime
import getpass # masked password input
import openpyxl  # A Python library to read/write Excel xlsx/xlsm files
from requests.auth import HTTPBasicAuth # Infoblox
from requests.packages.urllib3.exceptions import InsecureRequestWarning  # Infoblox
import configparser
import urllib3
import requests
from database import Database
from helper import gen_auth_cookie, enum

urllib3.disable_warnings(InsecureRequestWarning)  # Infoblox

config = configparser.ConfigParser()
config.read('config.ini')

#request_cookies = gen_auth_cookie(False)
#
path = config['local']['dc_server_file']
wb = openpyxl.load_workbook(path, read_only=False, keep_vba=True)  # definition of the excel workbook
ws = wb['all_DC_networks']  # definition of the excel workbook/worksheet (.active)
#
COLUMNS = enum(D=4, H=8, I=9)
# credentials
# passwd = getpass.getpass('Password :: ')
passwd = ''
user = 'dtcncat'
#
url_to_get_cookie = 'https://ipam03.wob.vw.vwg/wapi/v2.10/networkview'
response = requests.get(url_to_get_cookie ,auth=(user,passwd), verify=False)
authcookie = response.cookies['ibapauth']
#
request_cookies = {'ibapauth': authcookie} # set the cookie to a varibale
error_count = 0
netzrow = 2  # first row
max_netzrow = ws.max_row  # last row

db = Database()

def get_percentage(value):
    value_f = 0
    if value > 0:
        value_f = (value / 1000)  # devided by 10 - default is 10 - 1000 for %
    return value_f


# while loop until max rows for column A
while netzrow <= max_netzrow:
    cell_obj = ws.cell(row=netzrow, column=COLUMNS.D)  # read cell D2 - first request
    network = cell_obj.value  # define network with cell object (D2)

    try:
        url = config['ipam']['dhcp_util'].format(network)
        ans = requests.get(url, cookies=request_cookies, verify=False,
                           headers={'Content-Type': 'application/json'})  # reuse the cookie from first request

        networks = ans.json()  # assign the infoblox return values to variable - networks -

        #net_usage = 50 # CHANGE to 0
        #dhcp_usage = 50 # CHANGE to 0
        net_usage = get_percentage(networks[0]['utilization'])
        dhcp_usage = get_percentage(networks[0]['dhcp_utilization'])

        db.add((datetime.now(), network, net_usage, dhcp_usage))
        #db.add((datetime(2021, 10, 1), network, 30, 30))
        #db.add((datetime(2021, 9, 1), network, 30, 30))

        netzrow += 1  # +1 to use the next row until max_netzrow is reached

    except (NameError, IndexError, RuntimeError):
        netzrow += 1  # +1 to use the next row until max_netzrow is reached
        error_count += 1  # only count lines without content
        pass

print("An error occured for ", error_count, " rows")
wb.save(path)  # save workbook'''
